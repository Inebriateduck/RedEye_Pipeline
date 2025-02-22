# Note that you CANNOT have the target file open while running this script; doing so will result in an error

import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import datetime


def combine_and_process_csv_files():
    input_folder = r'XXXXXXXXX'  # Change this to your folder path

    # Regex pattern to match Unicode hex code form &#xNNNN;
    unicode_pattern = re.compile(r'&#x([0-9a-fA-F]+);')

    # Define a fill style for highlighting
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    combined_wb = Workbook()
    combined_ws = combined_wb.active
    combined_ws.title = "Combined Data"

    # Iterate over all CSV files in the input folder
    for file_name in os.listdir(input_folder):
        if file_name.endswith(".csv"):  # Process only CSV files
            file_path = os.path.join(input_folder, file_name)

            # Read CSV file as text (to preserve formatting)
            df = pd.read_csv(file_path, dtype=str)

            # Append rows from the CSV to the combined Excel sheet
            for _, row in df.iterrows():
                combined_ws.append(row.tolist())  # Convert row to list and append

    # Process the combined worksheet for Unicode hex codes
    for row in combined_ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                matches = unicode_pattern.findall(cell.value)
                if matches:
                    for match in matches:
                        # Replace hex code with the corresponding Unicode character
                        unicode_char = chr(int(match, 16))
                        cell.value = unicode_pattern.sub(unicode_char, cell.value, 1)
                    cell.fill = highlight_fill  # Highlight the cell

    folder_name = os.path.basename(input_folder)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_file = os.path.join(input_folder, f"{folder_name}_HBOutput_{timestamp}.xlsx")

    combined_wb.save(output_file)

    print(f"Processing complete. The output file has been saved as: {output_file}")


# Run the function
combine_and_process_csv_files()

#(C) Daniel Fry, 2025



